#!/usr/bin/python
""" hld2oit.py:

 Description: 	Tool intended to convert HLD format files to OIT format


 Input Parameters:
		HLD File: Location to the HLD excel file

 Output: OIT excel file

 Example:
		hld2oit.py "HLD_USC_AFF_vMCC_V.1.0.2.xls"

 Database:	N/A

 Created by : Daniel Jaramillo
 Creation Date: 04/01/2019
 Modified by:     Date:
 All rights(C) reserved to Teoco
"""

import sys
import os
import tokenize
import pandas as pd
from StringIO import StringIO
from openpyxl import load_workbook
from LoggerInit import LoggerInit
from oit_mapping import oit_mapping

def get_vars_divs(formula):
    """
    Description: Gets the variables and the divisors from a formula
    Input Parametes:
        formula
    """
    vars=[]
    divs=[]
    div=''
    start_div=False
    it=tokenize.generate_tokens(StringIO(formula).readline)
    for type,value,_,_,_ in it:
        if value=='/':
            if div:
                divs.append(div)
                div=''
            start_div=True
        elif value=='*':
            if div:
                divs.append(div)
                div=''
            start_div=False
        elif value==')':
            if div:
                divs.append(div)
                div=''
            start_div=False
        else:
            if type == 1:
                vars.append(value)
            if start_div and value !="(":
                div+=value
    vars=list(set(vars))
    divs=list(set(divs))
    return vars,divs

def create_tpt(kpi_name,formula,folder):
    """
    Description: Creates tpt function fomr a formula
    Input Parametes:
       kpi_name
       formula
       folder
    """
    app_logger=logger.get_logger("create_tpt")
    app_logger.info("Creating {kpi_name}={formula}"\
                    .format(kpi_name=kpi_name,formula=formula))
    vars,divs=get_vars_divs(formula)
    tpt_file_name='{schema}_TrolLocalFunctions.tpt'\
        .format(schema=metadata['Library Info']['SCHEMA'])
    with open(tpt_file_name,'a') as file:
        file.write('\n')
        file.write('@@PROTO\n')
        file.write('type=UF\n')
        file.write('id={kpi_name}\n'.format(kpi_name=kpi_name))
        file.write('location=Local.{folder}\n'.format(folder=folder))
        file.write('desc=\n')
        file.write('bitmap=\n')
        file.write('inpParamsNum={num_vars}\n'.format(num_vars=len(vars)))
        for idx,var in enumerate(vars):
            file.write('{idx}={var}, double, 1\n'.format(idx=idx+1,var=var))
        file.write('outParamsNum=1\n')
        file.write('1={kpi_name}, double\n'.format(kpi_name=kpi_name))
        file.write('keywordsNum=0\n')
        file.write('help=\n')
        file.write('@@CodeBegin\n')
        file.write('\n')
        for idx,var in enumerate(vars):
            vars[idx]='IsNull({var})'.format(var=var)
        vars_val="if ({vars}){{".format(vars='||'.join(vars))
        file.write(vars_val)
        file.write('\n')
        file.write('    {kpi_name} = NullDouble();\n'.format(kpi_name=kpi_name))
        file.write('    return true;\n')
        file.write('}\n')
        if divs:
            for idx,div in enumerate(divs):
                divs[idx]='{div} == 0'.format(div=div)
            divs_val="if ({divs}){{".format(divs='||'.join(divs))
            file.write(divs_val)
            file.write('\n')
            file.write('    {kpi_name} = 0;\n'.format(kpi_name=kpi_name))
            file.write('    return true;\n')
            file.write('}\n')
        file.write('{kpi_name}={formula};'\
                   .format(kpi_name=kpi_name,formula=formula))
        file.write('\n')
        file.write('return true;\n')
        file.write('\n')
        file.write('@@CodeEnd\n')
        file.write('@@PROTO_END\n')
        file.write('\n')


def create_functions():
    """
    Description: creates the functions for the KPI counters
    """
    app_logger=logger.get_logger("create_functions")
    app_logger.info("Creating functions")
    tpt_file_name='{schema}_TrolLocalFunctions.tpt'\
        .format(schema=metadata['Library Info']['SCHEMA'])
    #Make file emty
    open(tpt_file_name, 'w').close()
    df=metadata['Keys_Counters_KPIs'].dropna(subset=['KPI Formula'])
    for idx,kpi in df.iterrows():
        create_tpt(kpi['Counter/KPI DB Name'],
            kpi['KPI Formula'],
            metadata['Library Info']['SCHEMA'])

def parse_front_page(xl):
    """
    Description:  Parse the Front Page sheet
    Input Parametes:
        xl: Pandas excel file object
    """
    global metadata
    metadata['Front Page']={}
    app_logger=logger.get_logger("parse_front_page")
    app_logger.info("Parsing front page")
    df=xl.parse('Front Page')
    df=df.iloc[:,[0,1]].dropna(how='all')
    for index,row in df.iterrows():
        if row[0] == "Revision History":
            break
        metadata['Front Page'][row[0]]=row[1]

def parse_library_info(xl):
    """
    Description:  Parse the Library Info sheet
    Input Parametes:
        xl: Pandas excel file object
    """
    global metadata
    metadata['Library Info']={}
    app_logger=logger.get_logger("parse_library_info")
    app_logger.info("Parsing Library Info")
    df=xl.parse('Library Info')
    df=df.iloc[:,[1,2]].dropna(how='all')
    for index,row in df.iterrows():
        if row[0] == "Table Retention:":
            break
        metadata['Library Info'][row[0]]=row[1]


def parse_table(xl,sheet_name):
    """
    Description:  Parse the sheet in table format
    Input Parametes:
        xl: Pandas excel file object
        sheet name
    """
    global metadata
    metadata[sheet_name]={}
    app_logger=logger.get_logger("parse_table")
    app_logger.info("Parsing {sheet_name}".format(sheet_name=sheet_name))
    df=xl.parse(sheet_name)
    metadata[sheet_name]=df.iloc[2:,1:]


def load_hld(hld_file):
    """
    Description: Load the configuration from HLD file
    Input Parametes:
        hld_file: Excel containing the functional specification for the library
    """
    app_logger=logger.get_logger("load_hld "+hld_file)
    app_logger.info("Parsing HLD")
    xl=pd.ExcelFile(hld_file)
    parse_front_page(xl)
    parse_library_info(xl)
    parse_table(xl,"Entities")
    parse_table(xl,"Tables")
    parse_table(xl,"Keys_Counters_KPIs")



def write_oit():
    """
    Description: write to OIT
    Input Parametes:
        hld_file: Excel containing the functional specification for the library
    """
    app_logger=logger.get_logger("write_oit")
    app_logger.info("Creating OIT File")

    schema=metadata['Library Info']['SCHEMA']
    wb = load_workbook('template/EASY_PM_TEMPLATE_HELIX9.xlsx')

    #Populate Front Page
    for sheet,fields in oit_mapping.items():
        ws = wb[sheet]
        for field in fields:
            value=metadata[field['hld_sheet']][field['hld_field']]
            ws.cell(row=field['row'], column=field['column'], value=value)

    #Populate Entities related sheets
    ws_ent = wb['Entities']
    ws_cfg = wb['CFG Tables']
    ws_cfg_fields = wb['CFG Fields']
    for index,entity in metadata['Entities'].iterrows():
        #Populate Entities
        configuration_view=''
        if entity['Entity Type'] == 'Managed':
            configuration_view=entity['CFG Table or conf View']
        record=[entity['Entity Name'],
                entity['Element Alias'],
                entity['Parent Entity'],
                entity['Presentation'],
                configuration_view,
                entity['Universe']]
        ws_ent.append(record)
        #If entity is Managed we dont need to define conf views
        if entity['Entity Type'] == 'Managed':
            continue
        #Populate CFG Tables
        #Get table list for autopuplate
        df=metadata['Tables']
        df=df.loc[df['Entity'] == entity['Entity Name']].head()
        tables_arr=[]
        for index,table in df.iterrows():
            tables_arr.append(table['Table Name'])
        tables=','.join(tables_arr)
        configuration_view=entity['CFG Table or conf View'].split('.')[1]
        record=[configuration_view,entity['Entity Name'],tables]
        ws_cfg.append(record)
        #Populate CFG Fields
        key_list=entity['Keys'].split(',')
        for idx,key in enumerate(key_list):
            record=[entity['CFG Table or conf View'],
                    key,'VARCHAR2','YES',100,idx+1]
            ws_cfg_fields.append(record)

    #Populate Counter Sets
    ws_cs = wb['Counter Sets']
    ws_sum = wb['Summary Defn']
    for index,table in metadata['Tables'].dropna(how='all').iterrows():
        #Fill Counter Sets
        if table['Base Granularity'] == '5M':
            granularity=5
        elif table['Base Granularity'] == '15M':
            granularity=15
        elif table['Base Granularity'] == '30M':
            granularity=30
        elif table['Base Granularity'] == 'HR':
            granularity=60
        elif table['Base Granularity'] == 'DY':
            granularity=1440
        record=[table['Table Name'],
                table['Alias Table Name '],
                table['Counter Group in RD'],
                table['Entity'],
                'YES',
                granularity,
                table['Universe']
               ]
        ws_cs.append(record)
        #Fill Summary Defn
        summaries=table['Time Summary'].split(',')
        for summary in summaries:
            record=[table['Table Name'],summary]
            ws_sum.append(record)

    #Populate Loaded Counters
    ws = wb['Loaded Counters']
    df=metadata['Keys_Counters_KPIs'].dropna(how='all')
    prev_counter_set=''
    aggr_list=['AVG','SUM','MAX','MIN']
    for index,counter in df.iterrows():
        size=''
        if counter['TYPE'] in ['GPI','PI','OI']:
            size=100
        if prev_counter_set!=counter['Table Name']:
            order=1
        else:
            order+=1
        if counter['Time Aggregation'] not in aggr_list:
            aggr_formula='NULL'
        else:
            aggr_formula=counter['Time Aggregation']
        if counter['Hierarchy Summary'] not in aggr_list:
            ent_aggr_formula='NULL'
        else:
            ent_aggr_formula=counter['Hierarchy Summary']
        record=[counter['Table Name'],
                counter['Counter/KPI DB Name'],
                counter['Vendor Counter Name'],
                counter['Counter/KPI Display Name'],
                counter['TYPE'],
                counter['KPI Formula'],
                size,
                order,
                'YES',
                aggr_formula,
                aggr_formula,
                ent_aggr_formula,
                counter['Counter Description'],
                counter['Default Counter'],
                counter['Visible'],
                'YES',
        ]
        ws.append(record)
        prev_counter_set=counter['Table Name']
    wb.save("{schema}.xlsx".format(schema=schema))
    app_logger.info("{schema}.xlsx file created".format(schema=schema))

def main():
    app_logger=logger.get_logger("main")
    app_logger.info("Starting {script}".format(script=sys.argv[0]))
    #Validate the line arguments
    if len(sys.argv) < 2:
        app_logger.error("Usage {script} 'HLD File'"
                         .format(script=sys.argv[0]))
        app_logger.error("Example {script} 'HLD_USC_AFF_vMCC_V.1.0.2.xls'"
                         .format(script=sys.argv[0]))
        quit()

    hld_file=sys.argv[1]
    #Load configuration
    load_hld(hld_file)
    #Create OIT
    write_oit()
    #Create tpt functions
    create_functions()


#Application starts running here
if __name__ == "__main__":
    #If LOG_DIR environment var is not defined use /tmp as logdir
    if 'LOG_DIR' in os.environ:
        log_dir=os.environ['LOG_DIR']
    else:
        log_dir="/tmp"

    log_file=os.path.join(log_dir,"hld2oit.log")
    logger=LoggerInit(log_file,10)
    metadata={}
    main()
